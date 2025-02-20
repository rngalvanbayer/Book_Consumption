import modules
import credentials
import pandas as pd
import sqlalchemy
from credentials import dbxcat, dbxhost, dbxpath, dbxschema, dbxtoken
import os
import openpyxl
import encode




def main():
    # list all files in the sharepoint folder
    access_token = modules.get_graph_token()                            # Graph API token
    site_id = modules.get_site_id(access_token, 
                                 'bayergroup.sharepoint.com', 
                                 'BookConsumption')                     # sharepoint site id
    drive_id= modules.get_drive_id(access_token, site_id)               # sharepoint drive id
    folder_path = 'Encoding'                                       # child folder in sharepoint
    files = modules.get_sharepoint_files(access_token, 
                                         site_id, folder_path)          # list of all files in child folder
    df_files = files
    no_of_files = len(files)
    #no_of_files = 2
  
    
    # ======= start iteration here ========= 
    for i in range(no_of_files):
        filename = files[i]['name']
        if str('.pdf') in str(filename):
            print(filename)
            file_id = modules.get_file_id(access_token, site_id, drive_id, filename)
            modules.download_file(access_token, site_id, drive_id, file_id, filename)
            print("==> Downloading " + filename)

            # scan the file - finished data 
            #f_data = gptmodules.get_finished_data(filename, credentials.GPT4V_KEY, credentials.GPT4V_ENDPOINT)
            #finished_data = pd.concat([finished_data, f_data], ignore_index=True)
            #print(f_data)

            # scan the file - provided material
            #p_data = gptmodules.get_provided_data(filename, credentials.GPT4V_KEY, credentials.GPT4V_ENDPOINT)
            #provided_data = pd.concat([provided_data, p_data], ignore_index=True)
            #print(p_data)

 
            f_data, p_data = encode.encode(filename)
        
            # scan the filename - order item
            o_item = filename[3:]
            o_item = o_item[:-4]
            #print(o_item)

            # delete the pdf file and pngs
            modules.delete_file(filename)
            #f_data['Material No'] = str(order_item)         # order item removed
            merged_df = modules.alternate_rows(f_data, p_data)
            merged_df['Qty'] = merged_df['Qty'].str.replace(',', '').str.replace('.', '', regex=False).astype(int)
            #print(merged_df.to_markdown())
            fname = "GVT2-" + o_item + ".xlsx"
            merged_df.to_excel(fname, index=False, sheet_name=o_item)

    # ===== Databricks =====
    dbxcat = "efdataonelh_prd"
    dbxschema = "iaut_dmart"
    dbxhost     = credentials.dbxhost                                       ## Credentials
    dbxpath     = credentials.dbxpath                                      ## Credentials
    dbxtoken    = credentials.dbxtoken      

    extra_connect_args = {
        "_tls_verify_hostname": True,
        "_user_agent_entry": "ELSA Databricks Client",
    }


    engine = sqlalchemy.create_engine(
        f"databricks://token:{dbxtoken}@{dbxhost}?http_path={dbxpath}&catalog={dbxcat}&schema={dbxschema}",
        connect_args=extra_connect_args, echo=False,
    )

    cnxn = engine.connect()
    # === END Databricks

    # list all files in the directory
    files = os.listdir()

    # Convert the list of files to a DataFrame
    df = pd.DataFrame(files, columns=['Filename'])
    
    # filter out only GVT2 Excel files
    gvt2_df = df[df['Filename'].str.contains('GVT2-')]
  
    
    # loop around the numbers of GVT2 excel files
    for i in range(len(gvt2_df)):
        filename = gvt2_df.iloc[i]['Filename']
        #print(filename)
        print("==> Processing "+filename) 
        
        # load the excel file to a temporary dataframe
        df = pd.read_excel(filename)
        #print(df.to_markdown())
        
        #loop around the df:
        for j in range(len(df)):
            type = df.iloc[j]['Material Type']
            matno = str(df.iloc[j]['Material No'])
            batch = str(df.iloc[j]['Batch'])
            qty = df.iloc[j]['Qty']
            #print(type)
            
            
            # open excel file for reading
            workbook = openpyxl.load_workbook(filename=filename)
            sheet = workbook.active
            sheet['E1'] = 'Logged Qty' 
            
            # IF type is Bulk
            if type == 'Bulk':
                #print('bulk')
                sqlquery = "SELECT MATNR, CHARG, CINSM, CLABS, LFMON, LFGJA FROM efdataonelh_prd.generaldiscovery_matmgt_r.all_mchbh_view where MATNR LIKE '%" + matno + "' AND CHARG = '" + batch + "' AND LFGJA IN ('2024','2025') order by LFMON desc limit 1"
                #print(sqlquery)
                try:
                    df2 = pd.read_sql_query(sqlquery, cnxn)
                    #print(df2.to_markdown())
                    dbx_qty = (df2['CLABS'].sum()) * 1000
                    # write the quantity to the excel file
                    cellno = "E" + str(j+2)
                    sheet[cellno] = dbx_qty 
                    workbook.save(filename)
                    #print(dbx_qty)
                    #print(df2.to_markdown())
            
                except:
                    print("error")
            else:
                #print('finished goods')
                sqlquery = "SELECT MATNR, CHARG, CINSM, CLABS, LFMON, LFGJA FROM efdataonelh_prd.generaldiscovery_matmgt_r.all_mchbh_view where MATNR LIKE '%" + matno + "' AND CHARG = '" + batch + "' AND LFGJA IN ('2024','2025') order by LFMON desc limit 6"
                #print(sqlquery)
                try:
                    df2 = pd.read_sql_query(sqlquery, cnxn)
                    #print(df2.to_markdown())
                    dbx_qty = df2['CINSM'].sum()
                    if dbx_qty == 0:
                        dbx_qty = df2['CLABS'].sum()
                    # write the quantity to the excel file
                    cellno = "E" + str(j+2)
                    sheet[cellno] = dbx_qty
                    workbook.save(filename)
                    #print(dbx_qty)
                    #print(df2.to_markdown())
            
                except:
                    print("error")
    
    # Post-formatting excel        
    for i in range(len(gvt2_df)):
        filename = gvt2_df.iloc[i]['Filename']
        #print(filename)
        df = pd.read_excel(filename)
        workbook = openpyxl.load_workbook(filename=filename)
        sheet = workbook.active
        sheet['F1'] = 'diff' 
        sheet['G1'] = "% diff" 
        sheet['H1'] = 'Remarks'
        for cell in sheet[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        c = ['A1','B1','C1', 'E1', 'F1', 'G1', 'H1']
        thin = openpyxl.styles.Side(style='thin', color='000000')  # Black thin border
        border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
        for cr in c:
            cell = sheet[cr]
            cell.border = border
        sheet.column_dimensions['A'].width = 14
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 9
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 12
        sheet.column_dimensions['F'].width = 12
        sheet.column_dimensions['G'].width = 9
        sheet.column_dimensions['H'].width = 20
        workbook.save(filename)
        
        for j in range(len(df)):
            type = df.iloc[j]['Material Type']
            matno = str(df.iloc[j]['Material No'])
            batch = str(df.iloc[j]['Batch'])
            qty = df.iloc[j]['Qty']
            tqty = df.iloc[j]['Logged Qty']
            diff = abs(qty-tqty)           
            cellno = "F" + str(j+2)
            sheet[cellno] = diff 
            perc = diff / qty
            cellno = "G" + str(j+2)
            sheet[cellno] = perc 
            workbook.save(filename)
        
        for row in range(2, sheet.max_row + 1):
            cell = sheet[f'G{row}']
            cell.number_format = '0.00%'
            workbook.save(filename)
            
        df = pd.read_excel(filename)
        print(df.to_markdown())    
    
    # upload excel files to sharepoint        
    for i in range(len(gvt2_df)):
        filename = gvt2_df.iloc[i]['Filename']
        modules.upload_file(access_token, site_id, drive_id, 'Report', filename)
        print("==> Uploading " + filename)
        
    
    # delete pdf files in sharepoint
    for i in range(len(df_files)):
        filename = df_files[i]['name']
        if str('.pdf') in str(filename):
            file_id = modules.get_file_id(access_token, site_id, drive_id, filename)
            #modules.delete_file_sp(access_token, site_id, drive_id, file_id, filename)
            modules.delete_file_sp(access_token, drive_id, file_id)
            print("==> Delete " + filename)
    
    for i in range(len(gvt2_df)):
        filename = gvt2_df.iloc[i]['Filename']
        modules.delete_file(filename)
        print("==> Delete " + filename)
        

if __name__ == "__main__":
    main()